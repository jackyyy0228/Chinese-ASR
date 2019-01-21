from openpyxl import load_workbook
import re,os,sys
from number2chinese import *
from normalize_utils import *

def read_xlsx(file_name):
    wb = load_workbook(file_name)
    ws = wb.active
     
    rows = ws.rows
    columns = ws.columns
     
    content = []
    for idx,row in enumerate(rows):
        line = []
        for idx,col in enumerate(row):
            if idx == 0 and col.value is None:
                break
            if col.value is not None:
                line.append(col.value)
            else:
                break
        content.append(line)
    return content
'''
def read_xlsx(file_name):
    wb = load_workbook(file_name)
    ws = wb.active # 获取特定的 worksheet
     
    rows = ws.rows
    columns = ws.columns
     
    # 行迭代
    content = []
    for idx,row in enumerate(rows):
        line = [col.value for col in row] # ['No', '文章', '問題', '選項1', '選項2', '選項3', '選項4', '正確答案']
        content.append(line)
    
    return content
'''
def num_to_7digits(num):
    num = str(num)
    new = num
    for idx in range(7-len(num)):
        new = '0' + new
    return new

def merge_choice(choices):
    c = ''
    chin = ['一','二','三','四']
    for idx,choice in enumerate(choices):
        c += chin[idx] + ' ' + choice + ' '
    return c
if __name__ == '__main__':
    kaggle_dir = sys.argv[1]
    output_type = sys.argv[2]
    abc_type = sys.argv[3]
    words_file = sys.argv[4]

    xlxs_path = os.path.join(kaggle_dir,'answer.xlsx') 
    content = read_xlsx(xlxs_path)
    
    ## get all delete_symbols
    '''
    texts = ''
    for row in content:
        for idx in [1,2]:
            texts += row[idx]
    texts = strQ2B(texts)
    not_chinese = check_not_chinese(texts)
    print(not_chinese)
    exit()
    '''
    if output_type == 'text':
        word_list = get_word_list(words_file)
    for idx,row in enumerate(content):
        if idx == 0:
            continue
        No,passage,question,c1,c2,c3,c4 = row[:7]
        No = int(No)
        label = num_to_7digits(No)
        if output_type == 'utt2spk':
            for typ in ['A','B','C']:
                if typ == abc_type.upper():
                    print(typ+label,typ+label)
        elif output_type == 'wav.scp':
            for typ in ['A','B','C']:
                wav_path = os.path.join(kaggle_dir,'data/wav/',typ,typ+label+'.wav')
                wav_path = os.path.abspath(wav_path)
                if typ == abc_type.upper():
                    print(typ+label,wav_path)
        elif output_type == 'text':
            p = normalize(passage,word_list)
            q = normalize(question,word_list)
            c1 = normalize(str(c1),word_list)
            c2 = normalize(str(c2),word_list)
            c3 = normalize(str(c3),word_list)
            c4 = normalize(str(c4),word_list)
            c = merge_choice([c1,c2,c3,c4])
            for typ,text in [('A',p),('B',q),('C',c)]:
                if typ == abc_type.upper():
                    print(typ+label,text)
        
        
